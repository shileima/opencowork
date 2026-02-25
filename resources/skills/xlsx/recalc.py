#!/usr/bin/env python3
"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
from openpyxl import load_workbook


def _get_libreoffice_user_macro_dir() -> str:
    """Return LibreOffice user macro directory for the current OS."""
    system = platform.system()
    if system == 'Darwin':
        return os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    if system == 'Windows':
        appdata = os.environ.get('APPDATA') or os.path.join(os.path.expanduser('~'), 'AppData', 'Roaming')
        return os.path.join(appdata, 'LibreOffice', '4', 'user', 'basic', 'Standard')
    # Linux and others
    return os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')


def _which(cmd: str) -> str | None:
    """Minimal cross-platform 'which' for executables."""
    from shutil import which
    return which(cmd)


def _find_soffice() -> str:
    """Find LibreOffice 'soffice' executable; return command to execute."""
    found = _which('soffice')
    if found:
        return found

    if platform.system() == 'Windows':
        candidates = [
            os.path.join(os.environ.get('PROGRAMFILES', r'C:\\Program Files'), 'LibreOffice', 'program', 'soffice.exe'),
            os.path.join(os.environ.get('PROGRAMFILES(X86)', r'C:\\Program Files (x86)'), 'LibreOffice', 'program', 'soffice.exe'),
        ]
        for c in candidates:
            if os.path.exists(c):
                return c

    # macOS app bundle fallback
    if platform.system() == 'Darwin':
        mac = '/Applications/LibreOffice.app/Contents/MacOS/soffice'
        if os.path.exists(mac):
            return mac

    # Let subprocess raise a clear error
    return 'soffice'


def setup_libreoffice_macro():
    """Setup LibreOffice macro for recalculation if not already configured"""
    macro_dir = _get_libreoffice_user_macro_dir()
    
    macro_file = os.path.join(macro_dir, 'Module1.xba')
    
    if os.path.exists(macro_file):
        with open(macro_file, 'r') as f:
            if 'RecalculateAndSave' in f.read():
                return True
    
    if not os.path.exists(macro_dir):
        soffice = _find_soffice()
        subprocess.run([soffice, '--headless', '--terminate_after_init'],
                      capture_output=True, timeout=10)
        os.makedirs(macro_dir, exist_ok=True)
    
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''
    
    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """
    Recalculate formulas in Excel file and report any errors
    
    Args:
        filename: Path to Excel file
        timeout: Maximum time to wait for recalculation (seconds)
    
    Returns:
        dict with error locations and counts
    """
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}
    
    abs_path = str(Path(filename).absolute())
    
    if not setup_libreoffice_macro():
        return {'error': 'Failed to setup LibreOffice macro'}
    
    soffice = _find_soffice()
    cmd = [
        soffice, '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]
    
    # Handle timeout command differences between Linux and macOS
    if platform.system() != 'Windows':
        timeout_cmd = 'timeout' if platform.system() == 'Linux' else None
        if platform.system() == 'Darwin':
            # Check if gtimeout is available on macOS
            try:
                subprocess.run(['gtimeout', '--version'], capture_output=True, timeout=1, check=False)
                timeout_cmd = 'gtimeout'
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass

        if timeout_cmd:
            cmd = [timeout_cmd, str(timeout)] + cmd

    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if result.returncode != 0 and result.returncode != 124:  # 124 is timeout exit code
        error_msg = result.stderr or 'Unknown error during recalculation'
        if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
            return {'error': 'LibreOffice macro not configured properly'}
        else:
            return {'error': error_msg}
    
    # Check for Excel errors in the recalculated file - scan ALL cells
    try:
        wb = load_workbook(filename, data_only=True)
        
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check ALL rows and columns - no limits
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break
        
        wb.close()
        
        # Build result summary
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {}
        }
        
        # Add non-empty error categories
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]  # Show up to 20 locations
                }
        
        # Add formula count for context - also check ALL cells
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_formulas.close()
        
        result['total_formulas'] = formula_count
        
        return result
        
    except Exception as e:
        return {'error': str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)
    
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()
